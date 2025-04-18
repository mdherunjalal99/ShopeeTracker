"""
Excel Handler Module

This module handles all interactions with the Excel file, including
reading product links, writing prices, and calculating discounts.
"""

import logging
import string
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelHandler:
    """Class for handling Excel operations."""
    
    def __init__(self, filepath: str):
        """
        Initialize the Excel handler.
        
        Args:
            filepath: Path to the Excel file
        """
        self.filepath = filepath
        self.workbook = openpyxl.load_workbook(filepath)
        self.sheet = self.workbook.active
        
        # Load the data with pandas for easier data manipulation
        self.df = pd.read_excel(filepath)
        
        # Cache column letters
        self._column_cache = {}
    
    def column_index_to_letter(self, index: int) -> str:
        """
        Convert a column index to a column letter.
        
        Args:
            index: The column index (0-based)
            
        Returns:
            The column letter (A, B, C, etc.)
        """
        if index in self._column_cache:
            return self._column_cache[index]
        
        letter = get_column_letter(index + 1)
        self._column_cache[index] = letter
        return letter
    
    def column_letter_to_index(self, letter: str) -> int:
        """
        Convert a column letter to a column index.
        
        Args:
            letter: The column letter (A, B, C, etc.)
            
        Returns:
            The column index (0-based)
        """
        # Handle both single letters and letter combinations
        letter = letter.upper()
        result = 0
        for char in letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1
    
    def get_column_value(self, row: int, col: Union[str, int]) -> str:
        """
        Get the value from a specific cell.
        
        Args:
            row: The row number (1-based)
            col: The column letter or index
            
        Returns:
            The cell value as a string
        """
        if isinstance(col, int):
            col = self.column_index_to_letter(col)
        
        cell = self.sheet[f"{col}{row}"]
        return str(cell.value or "").strip()
    
    def get_product_links_and_variations(
        self, link_col: str, var1_col: str, var2_col: str
    ) -> Dict[int, Dict[str, str]]:
        """
        Extract product links and variations from the Excel file.
        
        Args:
            link_col: The column letter containing product links
            var1_col: The column letter containing the first variation
            var2_col: The column letter containing the second variation
            
        Returns:
            Dict mapping row indices to product information
        """
        products = {}
        
        # Start from row 2 (assuming row 1 is header)
        for row_idx, row in self.df.iterrows():
            # Add 2 to row_idx to match Excel row numbers (1-based indexing + header row)
            excel_row = row_idx + 2
            
            # Get the product link
            link = row.get(link_col, "").strip()
            if not link or "shopee" not in link.lower():
                continue
            
            # Get the variations
            var1 = row.get(var1_col, "").strip() if var1_col else ""
            var2 = row.get(var2_col, "").strip() if var2_col else ""
            
            products[excel_row] = {
                "link": link,
                "var1": var1,
                "var2": var2
            }
        
        return products
    
    def find_or_create_date_column(self, date_str: str) -> int:
        """
        Find or create a column for the given date.
        
        Args:
            date_str: The date string in YYYY-MM-DD format
            
        Returns:
            The column index
        """
        # Look for existing column with this date
        for col_idx, col_name in enumerate(self.df.columns):
            if col_name == date_str:
                return col_idx
        
        # Column doesn't exist, find the last column
        last_col_idx = len(self.df.columns)
        
        # Add the date column to the DataFrame
        self.df[date_str] = None
        
        # Update the Excel sheet with the new column header
        col_letter = self.column_index_to_letter(last_col_idx)
        self.sheet[f"{col_letter}1"] = date_str
        
        return last_col_idx
    
    def write_prices(self, results: List[Tuple[int, Optional[float]]], date_str: str):
        """
        Write prices to the Excel file.
        
        Args:
            results: List of tuples containing (row_index, price)
            date_str: The date string for the column header
        """
        # Find or create the date column
        date_col_idx = self.find_or_create_date_column(date_str)
        date_col_letter = self.column_index_to_letter(date_col_idx)
        
        # Update the prices in the Excel sheet
        for row_idx, price in results:
            if price is not None:
                # Update the DataFrame
                self.df.loc[row_idx - 2, date_str] = price
                
                # Update the Excel cell
                self.sheet[f"{date_col_letter}{row_idx}"] = price
    
    def calculate_discounts(self, discount_col: str):
        """
        Calculate discount percentages compared to average prices.
        
        Args:
            discount_col: The column letter for storing discount percentages
        """
        # Get the discount column index
        discount_col_idx = self.column_letter_to_index(discount_col)
        
        # Find the columns containing prices (date columns)
        date_col_indices = []
        for col_idx, col_name in enumerate(self.df.columns):
            try:
                # Check if the column name is a valid date
                datetime.strptime(str(col_name), "%Y-%m-%d")
                date_col_indices.append(col_idx)
            except (ValueError, TypeError):
                continue
        
        if not date_col_indices:
            logger.warning("No date columns found, cannot calculate discounts")
            return
        
        # Calculate average prices and discounts
        for row_idx, row in self.df.iterrows():
            # Skip rows without product links
            if row_idx < 2:  # Skip configuration rows
                continue
            
            # Extract prices from date columns
            prices = []
            for col_idx in date_col_indices:
                price = row.iloc[col_idx]
                if isinstance(price, (int, float)) and price > 0:
                    prices.append(price)
            
            if not prices:
                continue
            
            # Calculate average price
            avg_price = sum(prices) / len(prices)
            
            # Get the current price (most recent date)
            current_price = prices[-1] if prices else 0
            
            # Calculate discount percentage
            if avg_price > 0 and current_price > 0:
                discount_percent = (1 - (current_price / avg_price)) * 100
            else:
                discount_percent = 0
            
            # Update the discount column in the DataFrame
            discount_col_name = self.df.columns[discount_col_idx]
            self.df.loc[row_idx, discount_col_name] = discount_percent
            
            # Update the Excel cell
            excel_row = row_idx + 2  # Adjust for Excel row numbers
            self.sheet[f"{discount_col}{excel_row}"] = discount_percent
            
            # Format positive discounts in green, negative in red
            cell = self.sheet[f"{discount_col}{excel_row}"]
            if discount_percent > 0:
                cell.font = Font(color="00AA00")  # Green
            elif discount_percent < 0:
                cell.font = Font(color="AA0000")  # Red
    
    def save(self):
        """Save the Excel file."""
        self.workbook.save(self.filepath)
